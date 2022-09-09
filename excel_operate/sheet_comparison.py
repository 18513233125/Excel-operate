# -*- encoding: utf-8 -*-
'''
@File    :   sheet_comparison.py
@Author  :   北极星光 
@Contact :   light22@126.com
'''

from excel_operate import ExcelOperate
from sheet_copy import SheetCopy
from list_operate import *
from openpyxl.utils import get_column_letter


class SheetComparison:
    def __init__(self, src_excel, cmp_excel, report_path: str = None) -> None:
        '''===================================\n
        传入两个Excel对象，进行比较。生成报告
        src_excel: 原Excel对象，可传入ExcelOperate类对象
        cmp_excel: 待比较的Excel对象，可传入ExcelOperate类对象
        report_path: 报告结果保存路径
        '''
        self.src_excel = src_excel
        self.cmp_excel = cmp_excel
        self.report_path = report_path

    def set_title_row(self, src_title_row: int, cmp_title_row: int):
        '''===================================\n
        此方法可以为原工作表及待比较工作表设置表头行
        src_title_row: 原Excel工作表的表头行，可传入整数类型
        cmp_title_row: 待比较Excel工作表的表头行，可传入整数类型
        '''
        self.src_title_row = src_title_row
        self.cmp_title_row = cmp_title_row

    def set_key_col(self, src_key_col: int, cmp_key_col: int):
        '''===================================\n
        此方法可以为原工作表及待比较工作表设置关键列
        src_key_col: 原Excel工作表的关键列，可传入整数类型
        cmp_key_col: 待比较Excel工作表的关键列，可传入整数类型
        '''
        self.src_key_col = src_key_col
        self.cmp_key_col = cmp_key_col

    def compare(self):
        '''===================================\n
        对比工作表：将原工作表及目标工作表的表头行和关键列设置好之后即可使用此方法对比工作表，并生成对比报告。
        '''
        # 如果目标工作表的表头行与原工作表的表头行不在同一行
        if self.src_title_row != self.cmp_title_row:
            if self.src_title_row > self.cmp_title_row:
                self.cmp_excel.insert_rows(
                    0, self.src_title_row-self.cmp_title_row)
                self.cmp_title_row = self.src_title_row
            else:
                self.src_excel.insert_rows(
                    0, self.cmp_title_row-self.src_title_row)
                self.src_title_row = self.cmp_title_row

        # 如果目标工作表的关键列与原工作表的关键列不在同一列
        if self.src_key_col != self.cmp_key_col:
            if self.src_key_col > self.cmp_key_col:
                self.cmp_excel.insert_rows(
                    0, self.src_key_col-self.cmp_key_col)
                self.cmp_key_col = self.src_key_col
            else:
                self.src_excel.insert_rows(
                    0, self.cmp_key_col-self.src_key_col)
                self.src_key_col = self.cmp_key_col

        # 如果目标工作表的表头行与原工作表的表头行不相同
        src_title_list = [i.value for i in self.src_excel.ws[self.src_title_row]]
        cmp_title_list = [i.value for i in self.cmp_excel.ws[self.cmp_title_row]]
        if src_title_list != cmp_title_list:
            pass

        # 如果目标工作表的关键列与原工作表的关键列不相同
        src_key_list = [i.value for idx,i in enumerate(self.src_excel.ws[get_column_letter(self.src_key_col)],1) if idx>self.src_title_row]
        if self.src_excel.ws[get_column_letter(self.src_key_col)] != self.cmp_excel.ws[get_column_letter(self.cmp_key_col)]:
            pass


# 调试
if __name__ == '__main__':
    src_excel = ExcelOperate('tests\示例.xlsx')
    cmp_excel = ExcelOperate('tests\示例 - 对比.xlsx')
    report_path = 'D:/Desktop/对比报告.xlsx'
    cmper = SheetComparison(src_excel, cmp_excel, report_path)
    cmper.set_title_row(2, 2)
    cmper.set_key_col(2, 2)
